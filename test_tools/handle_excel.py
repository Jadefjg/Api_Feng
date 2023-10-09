from openpyxl import load_workbook

from test_tools.handle_log import Log
from test_tools.handle_path import case_data_dir
from test_tools.handle_ini import conf

import pprint


class HandleExcel:

    def __init__(self,file_name,sheet_name):
        Log.info(msg='用例文件名称:')
        Log.info(msg=file_name)
        Log.info(msg='sheet_name:')
        Log.info(msg=sheet_name)
        self.file_name = file_name      # excel 文件
        self.work_book = load_workbook(filename=self.file_name)  # 初始化表格对象实例属性
        self.sheet_name = sheet_name    # 初始化表格对象， 实例属性

    # # 获取 excel 测试用例数据
    # def get_test_case(self):
    #     wb = load_workbook(filename=self.file_name)              # 实例化
    #     sh = wb[self.sheet_name]                                 # 获取 sheetname
    #     all_excel_data = list(sh.iter_rows(values_only=True))    # 获取表格所有数据
    #     excel_title = all_excel_data[0]                          # 获取表头
    #     case_data_list = all_excel_data[1:]                      # 获取所有的用例数据
    #     test_case_list = []                                      # 测试用例数据
    #     for case in case_data_list:
    #         test_case = dict(zip(excel_title,case))              # 数据拼接，用表头与测试用例数据进行组装拼成 dict
    #         test_case_list.append(test_case)                     # 将每次拼接好的测试添加到test_case_list
    #     wb.close()
    #     return test_case_list

    # 获取表头和所有测试用例数据
    def __get_all_data(self):
        all_datas = list(self.sheet.iter_rows(values_only=True))
        # print(list(all_datas))
        title = all_datas[0] # 获取表头，list切片
        case_datas = all_datas[1:] # 获取测试用例数据，list切片
        Log.info(msg='excel读取到的测试用例数据：')
        Log.info(msg=case_datas)
        Log.info(msg='excel读取到的测试用例表头数据：')
        Log.info(msg=title)
        return title,case_datas

    # 数据拼接 (dict)
    def get_case_data_dict(self):
        case_list = []
        title, case_datas = self.__get_all_data()
        for val in case_datas:
            result = dict(zip(title, val))
            case_list.append(result)
        self.__close_excel()
        Log.info(msg='测试用例数据拼接完成：')
        Log.info(msg=case_list)
        return case_list

    # 保存测试结果到excel
    def write_result(self,rows,columns,result=None):
        self.sheet.cell(row=rows,column=columns).value=result
        self.__save_excel()
        self.__close_excel()

    # 保存excel
    def __save_excel(self):
        self.work_book.save(filename=self.file_name)

    # excel关闭
    def __close_excel(self):
        self.work_book.close()


if __name__ == '__main__':
    sheet_name = conf.get(section='sheet',option='name')
    cl = HandleExcel(file_path=case_data_dir,sheet_name=sheet_name)
    case_list = cl.get_case_data_dict()
    pprint.pprint(case_list)